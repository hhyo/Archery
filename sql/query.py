# -*- coding: UTF-8 -*-
import datetime
import logging
import re
import os
import shutil
import time
import traceback

import simplejson as json
import sqlparse
from openpyxl.workbook import Workbook
from wsgiref.util import FileWrapper
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import permission_required
from django.core import serializers
from django.db import connection
from django.db import transaction
from django.db.models import Q, Min
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse

from common.config import SysConfig
from common.utils.api import BASE_DIR, async
from common.utils.const import WorkflowDict
from common.utils.extend_json_encoder import ExtendJSONEncoder
from sql.utils.dao import Dao
from sql.utils.data_masking import Masking
from sql.utils.resource_group import user_instances, user_groups
from sql.utils.workflow import Workflow
from .models import Users, Instance, QueryPrivilegesApply, QueryPrivileges, QueryLog, ResourceGroup, ExportQuery
from sql.engines import get_engine
logger = logging.getLogger('default')

data_masking = Masking()
workflowOb = Workflow()


# 查询权限申请用于工作流审核回调
def query_audit_call_back(workflow_id, workflow_status):
    # 更新业务表状态
    apply_info = QueryPrivilegesApply()
    apply_info.apply_id = workflow_id
    apply_info.status = workflow_status
    apply_info.save(update_fields=['status'])
    # 审核通过插入权限信息，批量插入，减少性能消耗
    if workflow_status == WorkflowDict.workflow_status['audit_success']:
        apply_queryset = QueryPrivilegesApply.objects.get(apply_id=workflow_id)
        # 库权限
        if apply_queryset.priv_type == 1:
            insertlist = [QueryPrivileges(
                user_name=apply_queryset.user_name,
                user_display=apply_queryset.user_display,
                instance_name=apply_queryset.instance_name, db_name=db_name,
                table_name=apply_queryset.table_list, valid_date=apply_queryset.valid_date,
                limit_num=apply_queryset.limit_num, priv_type=apply_queryset.priv_type) for db_name in
                apply_queryset.db_list.split(',')]
        # 表权限
        elif apply_queryset.priv_type == 2:
            insertlist = [QueryPrivileges(
                user_name=apply_queryset.user_name,
                user_display=apply_queryset.user_display,
                instance_name=apply_queryset.instance_name, db_name=apply_queryset.db_list,
                table_name=table_name, valid_date=apply_queryset.valid_date,
                limit_num=apply_queryset.limit_num, priv_type=apply_queryset.priv_type) for table_name in
                apply_queryset.table_list.split(',')]
        QueryPrivileges.objects.bulk_create(insertlist)


# 查询权限校验
def query_priv_check(user, instance_name, db_name, sql_content, limit_num):
    result = {'status': 0, 'msg': 'ok', 'data': {'priv_check': 1, 'limit_num': 0}}
    instance = Instance.objects.get(instance_name=instance_name)
    # 检查用户是否有该数据库/表的查询权限
    if user.is_superuser:
        if SysConfig().sys_config.get('admin_query_limit', 5000):
            user_limit_num = int(SysConfig().sys_config.get('admin_query_limit'))
        else:
            user_limit_num = 5000
        limit_num = int(user_limit_num) if int(limit_num) == 0 else min(int(limit_num), int(user_limit_num))
        result['data']['limit_num'] = limit_num
        return result

    # 查看表结构和执行计划，inception会报错，故单独处理，explain直接跳过不做校验
    elif re.match(r"^show\s+create\s+table", sql_content.lower()):
        tb_name = re.sub('^show\s+create\s+table', '', sql_content, count=1, flags=0).strip()
        # 先判断是否有整库权限
        db_privileges = QueryPrivileges.objects.filter(user_name=user.username, instance_name=instance_name,
                                                       db_name=db_name, priv_type=1,
                                                       valid_date__gte=datetime.datetime.now(), is_deleted=0)
        # 无整库权限再验证表权限
        if len(db_privileges) == 0:
            tb_privileges = QueryPrivileges.objects.filter(user_name=user.username, instance_name=instance_name,
                                                           db_name=db_name, table_name=tb_name, priv_type=2,
                                                           valid_date__gte=datetime.datetime.now(), is_deleted=0)
            if len(tb_privileges) == 0:
                result['status'] = 1
                result['msg'] = '你无' + db_name + '.' + tb_name + '表的查询权限！请先到查询权限管理进行申请'
                return result
    # sql查询, 可以校验到表级权限
    elif instance.db_type == 'mysql':
        # 首先使用inception的语法树打印获取查询涉及的的表
        table_ref_result = data_masking.query_table_ref(sql_content + ';', instance_name, db_name)

        # 正确解析拿到表数据，可以校验表权限
        if table_ref_result['status'] == 0:
            table_ref = table_ref_result['data']
            # 获取表信息,校验是否拥有全部表查询权限
            QueryPrivilegesOb = QueryPrivileges.objects.filter(user_name=user.username, instance_name=instance_name)
            # 先判断是否有整库权限
            for table in table_ref:
                db_privileges = QueryPrivilegesOb.filter(db_name=table['db'], priv_type=1,
                                                         valid_date__gte=datetime.datetime.now(),
                                                         is_deleted=0)
                # 无整库权限再验证表权限
                if len(db_privileges) == 0:
                    tb_privileges = QueryPrivilegesOb.filter(db_name=table['db'], table_name=table['table'],
                                                             valid_date__gte=datetime.datetime.now(), is_deleted=0)
                    if len(tb_privileges) == 0:
                        result['status'] = 1
                        result['msg'] = '你无' + table['db'] + '.' + table['table'] + '表的查询权限！请先到查询权限管理进行申请'
                        return result

        # 获取表数据报错，检查配置文件是否允许继续执行，并进行库权限校验
        
    table_ref = None
    # 校验库权限，防止inception的语法树打印错误时连库权限也未做校验
    privileges = QueryPrivileges.objects.filter(user_name=user.username, instance_name=instance_name,
                                                db_name=db_name,
                                                valid_date__gte=datetime.datetime.now(),
                                                is_deleted=0)
    if len(privileges) == 0:
        result['status'] = 1
        result['msg'] = '你无' + db_name + '数据库的查询权限！请先到查询权限管理进行申请'
        return result
    if SysConfig().sys_config.get('query_check'):
        return table_ref_result
    else:
        result['data']['priv_check'] = 2

    # 获取查询涉及表的最小limit限制
    if table_ref:
        db_list = [table_info['db'] for table_info in table_ref]
        table_list = [table_info['table'] for table_info in table_ref]
        user_limit_num = QueryPrivileges.objects.filter(user_name=user.username,
                                                        instance_name=instance_name,
                                                        db_name__in=db_list,
                                                        table_name__in=table_list,
                                                        valid_date__gte=datetime.datetime.now(),
                                                        is_deleted=0).aggregate(Min('limit_num'))['limit_num__min']
        if user_limit_num is None:
            # 如果表没获取到则获取涉及库的最小limit限制
            user_limit_num = QueryPrivileges.objects.filter(user_name=user.username,
                                                            instance_name=instance_name,
                                                            db_name=db_name,
                                                            valid_date__gte=datetime.datetime.now(), is_deleted=0
                                                            ).aggregate(Min('limit_num'))['limit_num__min']
    else:
        # 如果表没获取到则获取涉及库的最小limit限制
        user_limit_num = QueryPrivileges.objects.filter(user_name=user.username,
                                                        instance_name=instance_name,
                                                        db_name=db_name,
                                                        valid_date__gte=datetime.datetime.now(),
                                                        is_deleted=0).aggregate(Min('limit_num'))['limit_num__min']
    limit_num = int(user_limit_num) if int(limit_num) == 0 else min(int(limit_num), int(user_limit_num))
    result['data']['limit_num'] = limit_num
    return result


# 获取查询权限申请列表
@permission_required('sql.menu_queryapplylist', raise_exception=True)
def getqueryapplylist(request):
    # 获取用户信息
    user = request.user

    limit = int(request.POST.get('limit'))
    offset = int(request.POST.get('offset'))
    limit = offset + limit
    search = request.POST.get('search', '')

    # 获取列表数据,申请人只能查看自己申请的数据,管理员可以看到全部数据,审核人可以看到自己审核的数据
    if user.is_superuser:
        lists = QueryPrivilegesApply.objects.all().filter(
            Q(title__contains=search) | Q(user_display__contains=search)).order_by('-apply_id')[
                offset:limit].values(
            'apply_id', 'title', 'instance_name', 'db_list', 'priv_type', 'table_list', 'limit_num', 'valid_date',
            'user_display', 'status', 'create_time', 'group_name'
        )
        count = QueryPrivilegesApply.objects.all().filter(title__contains=search).count()
    elif user.has_perm('sql.query_review'):
        # 先获取用户所在资源组列表
        group_list = user_groups(user)
        group_ids = [group.group_id for group in group_list]
        lists = QueryPrivilegesApply.objects.filter(group_id__in=group_ids).filter(
            Q(title__contains=search) | Q(user_display__contains=search)).order_by('-apply_id')[offset:limit].values(
            'apply_id', 'title', 'instance_name', 'db_list', 'priv_type', 'table_list', 'limit_num', 'valid_date',
            'user_display', 'status', 'create_time', 'group_name'
        )
        count = QueryPrivilegesApply.objects.filter(group_id__in=group_ids).filter(
            Q(title__contains=search) | Q(user_display__contains=search)).count()
    else:
        lists = QueryPrivilegesApply.objects.filter(user_name=user.username).filter(
            Q(title__contains=search) | Q(user_display__contains=search)).order_by('-apply_id')[offset:limit].values(
            'apply_id', 'title', 'instance_name', 'db_list', 'priv_type', 'table_list', 'limit_num', 'valid_date',
            'user_display', 'status', 'create_time', 'group_name'
        )
        count = QueryPrivilegesApply.objects.filter(user_name=user.username).filter(
            Q(title__contains=search) | Q(user_display__contains=search)).count()

    # QuerySet 序列化
    rows = [row for row in lists]

    result = {"total": count, "rows": rows}
    # 返回查询结果
    return HttpResponse(json.dumps(result, cls=ExtendJSONEncoder, bigint_as_string=True),
                        content_type='application/json')


# 申请查询权限
@permission_required('sql.query_applypriv', raise_exception=True)
def applyforprivileges(request):
    title = request.POST['title']
    instance_name = request.POST['instance_name']
    group_name = request.POST['group_name']
    group_id = ResourceGroup.objects.get(group_name=group_name).group_id
    priv_type = request.POST['priv_type']
    db_name = request.POST['db_name']
    valid_date = request.POST['valid_date']
    limit_num = request.POST['limit_num']

    # 获取用户信息
    user = request.user

    # 服务端参数校验
    result = {'status': 0, 'msg': 'ok', 'data': []}
    if int(priv_type) == 1:
        db_list = request.POST['db_list']
        if title is None or instance_name is None or db_list is None or valid_date is None or limit_num is None:
            result['status'] = 1
            result['msg'] = '请填写完整'
            return HttpResponse(json.dumps(result), content_type='application/json')
    elif int(priv_type) == 2:
        table_list = request.POST['table_list']
        if title is None or instance_name is None or db_name is None or valid_date is None or table_list is None or limit_num is None:
            result['status'] = 1
            result['msg'] = '请填写完整'
            return HttpResponse(json.dumps(result), content_type='application/json')
    try:
        user_instances(request.user, 'slave').get(instance_name=instance_name)
    except Exception:
        context = {'errMsg': '你所在组未关联该实例！'}
        return render(request, 'error.html', context)

    # 判断是否需要限制到表级别的权限
    # 库权限
    if int(priv_type) == 1:
        db_list = db_list.split(',')
        # 检查申请账号是否已拥整个库的查询权限
        own_dbs = QueryPrivileges.objects.filter(instance_name=instance_name, user_name=user.username,
                                                 db_name__in=db_list,
                                                 valid_date__gte=datetime.datetime.now(), priv_type=1,
                                                 is_deleted=0).values('db_name')
        own_db_list = [table_info['db_name'] for table_info in own_dbs]
        if own_db_list is None:
            pass
        else:
            for db_name in db_list:
                if db_name in own_db_list:
                    result['status'] = 1
                    result['msg'] = '你已拥有' + instance_name + '实例' + db_name + '库的全部查询权限，不能重复申请'
                    return HttpResponse(json.dumps(result), content_type='application/json')
    # 表权限
    elif int(priv_type) == 2:
        table_list = table_list.split(',')
        # 检查申请账号是否已拥有该表的查询权限
        own_tables = QueryPrivileges.objects.filter(instance_name=instance_name, user_name=user.username,
                                                    db_name=db_name,
                                                    table_name__in=table_list, valid_date__gte=datetime.datetime.now(),
                                                    priv_type=2, is_deleted=0).values('table_name')
        own_table_list = [table_info['table_name'] for table_info in own_tables]
        if own_table_list is None:
            pass
        else:
            for table_name in table_list:
                if table_name in own_table_list:
                    result['status'] = 1
                    result['msg'] = '你已拥有' + instance_name + '实例' + db_name + '.' + table_name + '表的查询权限，不能重复申请'
                    return HttpResponse(json.dumps(result), content_type='application/json')

    # 使用事务保持数据一致性
    try:
        with transaction.atomic():
            # 保存申请信息到数据库
            applyinfo = QueryPrivilegesApply()
            applyinfo.title = title
            applyinfo.group_id = group_id
            applyinfo.group_name = group_name
            applyinfo.audit_auth_groups = Workflow.audit_settings(group_id, WorkflowDict.workflow_type['query'])
            applyinfo.user_name = user.username
            applyinfo.user_display = user.display
            applyinfo.instance_name = instance_name
            if int(priv_type) == 1:
                applyinfo.db_list = ','.join(db_list)
                applyinfo.table_list = ''
            elif int(priv_type) == 2:
                applyinfo.db_list = db_name
                applyinfo.table_list = ','.join(table_list)
            applyinfo.priv_type = int(priv_type)
            applyinfo.valid_date = valid_date
            applyinfo.status = WorkflowDict.workflow_status['audit_wait']  # 待审核
            applyinfo.limit_num = limit_num
            applyinfo.create_user = user.username
            applyinfo.save()
            apply_id = applyinfo.apply_id

            # 调用工作流插入审核信息,查询权限申请workflow_type=1
            audit_result = workflowOb.addworkflowaudit(request, WorkflowDict.workflow_type['query'], apply_id)
            if audit_result['status'] == 0:
                # 更新业务表审核状态,判断是否插入权限信息
                query_audit_call_back(apply_id, audit_result['data']['workflow_status'])
    except Exception as msg:
        logger.error(traceback.format_exc())
        result['status'] = 1
        result['msg'] = str(msg)
    else:
        result = audit_result
    return HttpResponse(json.dumps(result), content_type='application/json')


# 用户的查询权限管理
def getuserprivileges(request):
    user_name = request.POST.get('user_name')
    limit = int(request.POST.get('limit'))
    offset = int(request.POST.get('offset'))
    limit = offset + limit
    search = request.POST.get('search', '')

    # 判断权限，除了管理员外其他人只能查看自己的权限信息，
    user = request.user

    # 获取用户的权限数据
    if user.is_superuser:
        if user_name != 'all':
            privileges_list = QueryPrivileges.objects.all().filter(user_name=user_name,
                                                                   is_deleted=0,
                                                                   table_name__contains=search,
                                                                   valid_date__gte=datetime.datetime.now()
                                                                   ).order_by('-privilege_id')[offset:limit]
            privileges_count = QueryPrivileges.objects.all().filter(user_name=user_name,
                                                                    is_deleted=0,
                                                                    table_name__contains=search,
                                                                    valid_date__gte=datetime.datetime.now()).count()
        else:
            privileges_list = QueryPrivileges.objects.all().filter(is_deleted=0,
                                                                   table_name__contains=search,
                                                                   valid_date__gte=datetime.datetime.now()
                                                                   ).order_by('-privilege_id')[offset:limit]
            privileges_count = QueryPrivileges.objects.all().filter(is_deleted=0,
                                                                    table_name__contains=search,
                                                                    valid_date__gte=datetime.datetime.now()
                                                                    ).count()
    else:
        privileges_list = QueryPrivileges.objects.filter(user_name=user.username,
                                                         table_name__contains=search,
                                                         is_deleted=0,
                                                         valid_date__gte=datetime.datetime.now()
                                                         ).order_by('-privilege_id')[offset:limit]
        privileges_count = QueryPrivileges.objects.filter(user_name=user.username,
                                                          table_name__contains=search,
                                                          is_deleted=0,
                                                          valid_date__gte=datetime.datetime.now()
                                                          ).count()

    # QuerySet 序列化
    privileges_list = serializers.serialize("json", privileges_list)
    privileges_list = json.loads(privileges_list)
    privilegeslist_result = []
    for i in range(len(privileges_list)):
        privileges_list[i]['fields']['id'] = privileges_list[i]['pk']
        privilegeslist_result.append(privileges_list[i]['fields'])

    result = {"total": privileges_count, "rows": privilegeslist_result}
    # 返回查询结果
    return HttpResponse(json.dumps(result), content_type='application/json')


# 变更权限信息
@permission_required('sql.query_mgtpriv', raise_exception=True)
def modifyqueryprivileges(request):
    privilege_id = request.POST.get('privilege_id')
    type = request.POST.get('type')
    result = {'status': 0, 'msg': 'ok', 'data': []}

    # type=1删除权限,type=2变更权限
    privileges = QueryPrivileges()
    if int(type) == 1:
        # 删除权限
        privileges.privilege_id = int(privilege_id)
        privileges.is_deleted = 1
        privileges.save(update_fields=['is_deleted'])
        return HttpResponse(json.dumps(result), content_type='application/json')
    elif int(type) == 2:
        # 变更权限
        valid_date = request.POST.get('valid_date')
        limit_num = request.POST.get('limit_num')
        privileges.privilege_id = int(privilege_id)
        privileges.valid_date = valid_date
        privileges.limit_num = limit_num
        privileges.save(update_fields=['valid_date', 'limit_num'])
        return HttpResponse(json.dumps(result), content_type='application/json')


# 查询权限审核
@permission_required('sql.query_review', raise_exception=True)
def queryprivaudit(request):
    # 获取用户信息
    user = request.user
    apply_id = int(request.POST['apply_id'])
    audit_status = int(request.POST['audit_status'])
    audit_remark = request.POST.get('audit_remark')

    if audit_remark is None:
        audit_remark = ''

    if Workflow.can_review(request.user, apply_id, 1) is False:
        context = {'errMsg': '你无权操作当前工单！'}
        return render(request, 'error.html', context)

    # 使用事务保持数据一致性
    try:
        with transaction.atomic():
            # 获取audit_id
            audit_id = Workflow.audit_info_by_workflow_id(workflow_id=apply_id,
                                                          workflow_type=WorkflowDict.workflow_type['query']).audit_id

            # 调用工作流接口审核
            audit_result = workflowOb.auditworkflow(request, audit_id, audit_status, user.username, audit_remark)

            # 按照审核结果更新业务表审核状态
            audit_detail = Workflow.audit_detail(audit_id)
            if audit_detail.workflow_type == WorkflowDict.workflow_type['query']:
                # 更新业务表审核状态,插入权限信息
                query_audit_call_back(audit_detail.workflow_id, audit_result['data']['workflow_status'])

    except Exception as msg:
        logger.error(traceback.format_exc())
        context = {'errMsg': msg}
        return render(request, 'error.html', context)

    return HttpResponseRedirect(reverse('sql:queryapplydetail', args=(apply_id,)))


# 获取SQL查询结果
@permission_required('sql.query_submit', raise_exception=True)
def query(request):
    instance_name = request.POST.get('instance_name')
    sql_content = request.POST.get('sql_content')
    db_name = request.POST.get('db_name')
    limit_num = request.POST.get('limit_num')

    result = {'status': 0, 'msg': 'ok', 'data': {}}
    try:
        instance = Instance.objects.get(instance_name=instance_name)
    except Instance.DoesNotExist:
        result['status'] = 1
        result['msg'] = '实例不存在'
        return result
    # 服务器端参数验证
    if sql_content is None or db_name is None or instance_name is None or limit_num is None:
        result['status'] = 1
        result['msg'] = '页面提交参数可能为空'
        return HttpResponse(json.dumps(result), content_type='application/json')

    sql_content = sql_content.strip()
    archer_config = SysConfig()
    if archer_config.get('disable_star'):
        if '*' in sql_content:
            result['status'] = 1
            result['msg'] = '不允许 * 标记, 请指定具体字段名.'
            return HttpResponse(json.dumps(result), content_type='application/json')
    # 获取用户信息
    user = request.user

    # 过滤注释语句和非查询的语句
    sql_content = ''.join(
        map(lambda x: re.compile(r'(^--\s+.*|^/\*.*\*/;\s*$)').sub('', x, count=1),
            sql_content.splitlines(1))).strip()
    # 去除空行
    sql_content = re.sub('[\r\n\f]{2,}', '\n', sql_content)

    sql_list = sql_content.strip().split('\n')
    for sql in sql_list:
        if re.match(r"^select|^show|^explain", sql.lower()):
            break
        else:
            result['status'] = 1
            result['msg'] = '仅支持^select|^show|^explain语法，请联系管理员！'
            return HttpResponse(json.dumps(result), content_type='application/json')

    # 执行第一条有效sql
    sql_content = sqlparse.split(sql_content)[0].rstrip(';')

    try:
        # 查询权限校验
        priv_check_info = query_priv_check(user, instance_name, db_name, sql_content, limit_num)

        if priv_check_info['status'] == 0:
            limit_num = priv_check_info['data']['limit_num']
            priv_check = priv_check_info['data']['priv_check']
        else:
            return HttpResponse(json.dumps(priv_check_info), content_type='application/json')

        if re.match(r"^explain", sql_content.lower()):
            limit_num = 0
        query_engine = get_engine(instance=instance)
        filter_result = query_engine.query_check(db_name=db_name, sql=sql_content, limit_num=limit_num)
        if filter_result.get('bad_query'):
            result['status'] = 1
            result['msg'] = filter_result.get('msg')
            return HttpResponse(json.dumps(result), content_type='application/json')
        else:
            sql_content = filter_result['filtered_sql']
        sql_content = sql_content + ';'

        # 执行查询语句,统计执行时间
        t_start = time.time()
        query_result = query_engine.query(db_name=str(db_name), sql=sql_content, limit_num=limit_num)
        t_end = time.time()
        query_result.query_time = "%5s" % "{:.4f}".format(t_end - t_start)

        # 数据脱敏，同样需要检查配置，是否开启脱敏，语法树解析是否允许出错继续执行
        hit_rule = 0 if re.match(r"^select", sql_content.lower()) else 2  # 查询是否命中脱敏规则，0, '未知', 1, '命中', 2, '未命中'
        masking = 2  # 查询结果是否正常脱敏，1, '是', 2, '否'
        t_start = time.time()
        # 仅对查询语句进行脱敏

        if SysConfig().sys_config.get('data_masking') and re.match(r"^select", sql_content.lower()):
            try:
                masking_result = query_engine.query_masking(db_name=db_name, sql=sql_content, resultset=query_result)
                if SysConfig().sys_config.get('query_check') and query_result.is_critical == True:
                    return HttpResponse(json.dumps(masking_result), content_type='application/json')
                else:
                    # 实际未命中, 则显示为未做脱敏
                    if query_result.is_masked:
                        masking = 1
                        hit_rule = 1
            except Exception:
                logger.error(traceback.format_exc())
                # 报错, 未脱敏, 未命中
                hit_rule = 2
                masking = 2
                if SysConfig().sys_config.get('query_check'):
                    result['status'] = 1
                    result['msg'] = '脱敏数据报错,请联系管理员'
                    return HttpResponse(json.dumps(result), content_type='application/json')

        t_end = time.time()
        query_result.mask_time = "%5s" % "{:.4f}".format(t_end - t_start)
        sql_result = query_result.__dict__
        sql_result['masking_cost_time'] = query_result.mask_time
        sql_result['cost_time'] = query_result.query_time

        result['data'] = sql_result

        # 成功的查询语句记录存入数据库
        if sql_result.get('error'):
            pass
        else:
            query_log = QueryLog()
            query_log.username = user.username
            query_log.user_display = user.display
            query_log.db_name = db_name
            query_log.instance_name = instance_name
            query_log.sqllog = sql_content
            if int(limit_num) == 0:
                limit_num = int(sql_result['affected_rows'])
            else:
                limit_num = min(int(limit_num), int(sql_result['affected_rows']))
            query_log.effect_row = limit_num
            query_log.cost_time = query_result.query_time
            query_log.priv_check = priv_check
            query_log.hit_rule = hit_rule
            query_log.masking = masking
            # 防止查询超时
            try:
                query_log.save()
            except:
                connection.close()
                query_log.save()
    except Exception as e:
        logger.error(traceback.format_exc())
        result['status'] = 1
        result['msg'] = str(e)

    # 返回查询结果
    try:
        return HttpResponse(json.dumps(result, cls=ExtendJSONEncoder, bigint_as_string=True),
                            content_type='application/json')
    except Exception:
        return HttpResponse(json.dumps(result, default=str, bigint_as_string=True, encoding='latin1'),
                            content_type='application/json')


# 获取sql查询记录
@permission_required('sql.menu_sqlquery', raise_exception=True)
def querylog(request):
    # 获取用户信息
    user = request.user

    limit = int(request.POST.get('limit'))
    offset = int(request.POST.get('offset'))
    limit = offset + limit
    search = request.POST.get('search', '')

    # 查询个人记录，超管查看所有数据
    if user.is_superuser:
        sql_log_count = QueryLog.objects.all().filter(
            Q(sqllog__contains=search) | Q(user_display__contains=search)).count()
        sql_log_list = QueryLog.objects.all().filter(
            Q(sqllog__contains=search) | Q(user_display__contains=search)).order_by(
            '-id')[offset:limit]
    else:
        sql_log_count = QueryLog.objects.filter(username=user.username).filter(sqllog__contains=search).count()
        sql_log_list = QueryLog.objects.filter(username=user.username).filter(sqllog__contains=search).order_by('-id')[
                       offset:limit]

    # QuerySet 序列化
    sql_log_list = serializers.serialize("json", sql_log_list)
    sql_log_list = json.loads(sql_log_list)
    sql_log = [log_info['fields'] for log_info in sql_log_list]

    result = {"total": sql_log_count, "rows": sql_log}
    # 返回查询结果
    return HttpResponse(json.dumps(result), content_type='application/json')


# 获取SQL执行计划
@permission_required('sql.optimize_sqladvisor', raise_exception=True)
def explain(request):
    sql_content = request.POST.get('sql_content')
    instance_name = request.POST.get('instance_name')
    db_name = request.POST.get('db_name')
    result = {'status': 0, 'msg': 'ok', 'data': []}

    # 服务器端参数验证
    if sql_content is None or instance_name is None:
        result['status'] = 1
        result['msg'] = '页面提交参数可能为空'
        return HttpResponse(json.dumps(result), content_type='application/json')

    sql_content = sql_content.strip()

    # 过滤非查询的语句
    if re.match(r"^explain", sql_content.lower()):
        pass
    else:
        result['status'] = 1
        result['msg'] = '仅支持explain开头的语句，请检查'
        return HttpResponse(json.dumps(result), content_type='application/json')

    # 执行第一条有效sql
    sql_content = sqlparse.split(sql_content)[0].rstrip(';')

    # 执行获取执行计划语句
    sql_result = Dao(instance_name=instance_name).mysql_query(str(db_name), sql_content)

    result['data'] = sql_result

    # 返回查询结果
    return HttpResponse(json.dumps(result, cls=ExtendJSONEncoder, bigint_as_string=True),
                        content_type='application/json')


# 获取SQL查询结果
@csrf_exempt
@permission_required('sql.query_submit', raise_exception=True)
def add_async_query(request):
    instance_name = request.POST.get('instance_name')
    sqlContent = request.POST.get('sql_content')
    dbName = request.POST.get('db_name')
    limit_num = request.POST.get('limit_num')
    auditor = request.POST.get('auditor')

    finalResult = {'status': 0, 'msg': 'ok', 'data': {}}

    # 服务器端参数验证
    if sqlContent is None or dbName is None or instance_name is None or limit_num is None:
        finalResult['status'] = 1
        finalResult['msg'] = '页面提交参数可能为空'
        return HttpResponse(json.dumps(finalResult), content_type='application/json')

    sqlContent = sqlContent.strip()
    if sqlContent[-1] != ";":
        finalResult['status'] = 1
        finalResult['msg'] = 'SQL语句结尾没有以;结尾，请重新修改并提交！'
        return HttpResponse(json.dumps(finalResult), content_type='application/json')

    # 获取用户信息
    user = request.user

    # 过滤注释语句和非查询的语句
    sqlContent = ''.join(
        map(lambda x: re.compile(r'(^--\s+.*|^/\*.*\*/;\s*$)').sub('', x, count=1),
            sqlContent.splitlines(1))).strip()
    # 去除空行
    sqlContent = re.sub('[\r\n\f]{2,}', '\n', sqlContent)

    sql_list = sqlContent.strip().split('\n')
    for sql in sql_list:
        if re.match(r"^select|^show|^explain", sql.lower()):
            break
        else:
            finalResult['status'] = 1
            finalResult['msg'] = '仅支持^select|^show|^explain语法，请联系管理员！'
            return HttpResponse(json.dumps(finalResult), content_type='application/json')

    # 取出该实例的连接方式,查询只读账号,按照分号截取第一条有效sql执行
    slave_info = Instance.objects.get(instance_name=instance_name)
    sqlContent = sqlContent.strip().split(';')[0]

    # 查询权限校验
    priv_check_info = query_priv_check(user, instance_name, dbName, sqlContent, limit_num)

    if priv_check_info['status'] == 0:
        pass
    else:
        return HttpResponse(json.dumps(priv_check_info), content_type='application/json')

    if re.match(r"^explain", sqlContent.lower()):
        limit_num = 0

    # 对查询sql增加limit限制
    if re.match(r"^select", sqlContent.lower()) and int(limit_num) > 0:
        if re.search(r"limit\s+(\d+)$", sqlContent.lower()) is None:
            if re.search(r"limit\s+\d+\s*,\s*(\d+)$", sqlContent.lower()) is None:
                sqlContent = sqlContent + ' limit ' + str(limit_num)

    sqlContent = sqlContent + ';'

    # 查询语句记录存入数据库
    query_log = QueryLog.objects.create(username=user.username, user_display=user.display, db_name=dbName,
                                        instance_name=instance_name, sqllog=sqlContent, effect_row=0)

    qe = ExportQuery.objects.create(query_log=query_log, auditor=Users.objects.get(username=auditor), status=0)

    do_async_query(request, qe, instance_name, dbName, slave_info.db_type, sqlContent, limit_num)
    finalResult['msg'] = '任务提交成功！后台拼命跑数据中... 请耐心等待钉钉或邮件通知！'

    return HttpResponse(json.dumps(finalResult, cls=ExtendJSONEncoder, bigint_as_string=True),
                        content_type='application/json')


@async
def do_async_query(request, export_query, instance_name, db_name, db_type, sql_content, limit_num):
    query_log = export_query.query_log
    instance = Instance.objects.get(instance_name=instance_name)
    query_engine = get_engine(instance=instance)
    filter_result = query_engine.query_check(db_name=db_name, sql=sql_content, limit_num=limit_num)
    if filter_result.get('bad_query'):
        result = {'status': 1, 'msg': filter_result.get('msg')}
        return HttpResponse(json.dumps(result), content_type='application/json')
    else:
        sql_content = filter_result['filtered_sql']
    sql_content = sql_content + ';'

    t_start = int(time.time())
    sql_result = query_engine.query(db_name=db_name, sql=sql_content, limit_num=limit_num)
    t_end = int(time.time())
    query_log.cost_time = t_end - t_start
    query_log.effect_row = sql_result["effect_row"]
    query_log.save()

    def write_result_to_excel(sr):
        try:
            file_dir = SysConfig().sys_config.get('query_result_dir', BASE_DIR)
            time_suffix = datetime.datetime.now().strftime("%m%d%H%M%S")
            file_name = '{}-{}-{}-{}'.format(export_query.query_log.username, query_log.instance_name, db_name, time_suffix)
            template_file = os.path.join(file_dir, file_name)

            workbook = Workbook(encoding='utf-8')
            ws = workbook.create_sheet('Sheet1')
            # 写入字段信息
            ws.append(sr["column_list"])

            # 写入数据段信息
            # for row in range(1, int(sr["effect_row"]) + 1):
            #     for col in range(0, len(sr["column_list"])):
            #         print(type(sr["rows"][row - 1][col]), sr["rows"][row - 1][col])
            #         value = '' if sr["rows"][row - 1][col] is None else sr["rows"][row - 1][col]
            #         ws.cell(row=row, column=col).value = value
            for row in range(len(sr["rows"])):
                ws.append(sr["rows"][row])
            workbook.save(template_file)
        except Exception as e:
            export_query.error_msg = str(traceback.print_exc())
            export_query.save(update_fields=['error_msg'])
            return str(e)
        return template_file

    try:
        file_path = ''
        if SysConfig().sys_config.get('data_masking'):
            if db_type == "mysql":
                # 仅对查询语句进行脱敏
                if re.match(r"^select", sql_content.lower()):
                    try:
                        masking_result = data_masking.data_masking(instance_name, db_name, sql_content, sql_result)
                    except Exception as e:
                        if SysConfig().sys_config.get('query_check'):
                            export_query.status = 1
                            export_query.error_msg = '脱敏数据报错,请联系管理员。报错：%s' % str(e)
                    else:
                        if masking_result['status'] == 0 or not SysConfig().sys_config.get('query_check'):
                            file_path = write_result_to_excel(sql_result)
                            export_query.status = 2
            else:
                file_path = write_result_to_excel(sql_result)
                export_query.status = 2
        else:
            file_path = write_result_to_excel(sql_result)
            export_query.status = 2
        export_query.result_file = file_path
    except Exception as e:
        export_query.error_msg = str(e)
        export_query.status = 1
    export_query.save()

    # 通知审核人审核
    audit_url = "{}://{}/export_query/".format(request.scheme, request.get_host())
    msg_content = '''导出查询（提取大量数据）下载申请等待您审批：\n发起人：{}\n实例名称：{}\n数据库：{}\n执行的sql查询：{}\n提取条数：{}\n操作时间：{}\n审批地址：{}\n'''.\
        format(query_log.user_display, query_log.instance_name, query_log.db_name, query_log.sqllog,
               query_log.effect_row, query_log.create_time, audit_url)
    from sql.utils.ding_api import DingSender
    DingSender().send_msg(export_query.auditor.ding_user_id, msg_content)


# 获取sql查询记录
@csrf_exempt
@permission_required('sql.query_submit', raise_exception=True)
def export_query_result(request):
    export_query_id = request.GET.get("id")
    qe = ExportQuery.objects.get(id=export_query_id)

    if qe.result_file:
        wrapper = FileWrapper(open(qe.result_file, "rb"))
        response = HttpResponse(wrapper, content_type='application/vnd.ms-excel')
        response['Content-Length'] = os.path.getsize(qe.result_file)
        response['Content-Disposition'] = 'attachment; filename="result.xls"'
        return response
    else:
        return HttpResponse(qe.error_msg)


# 获取sql查询记录
@csrf_exempt
@permission_required('sql.query_submit', raise_exception=True)
def export_query_audit(request):
    export_query_id = request.POST.get("id")
    is_allow = request.POST.get("is_allow", '')
    audit_msg = request.POST.get("audit_msg", '')
    qe = ExportQuery.objects.get(id=export_query_id)
    ql = qe.query_log
    applicant = Users.objects.get(username=ql.username)
    if qe.status == 1:
        # 执行失败
        msg = qe.error_msg
    elif qe.status == 2:
        # 审核
        user = request.user
        if not user.has_perm('sql.export_query_review'):
            msg = "你没有审核权限！"
        else:
            if is_allow == "yes":
                qe.status = 3
                msg = "已通过！"
                if os.path.exists(qe.result_file):
                    if os.path.getsize(qe.result_file) > 0:
                        shutil.copy2(qe.result_file, "{}.xls".format(qe.result_file.split('/')[-1]))
            else:
                qe.status = 4
                msg = "已拒绝！"
            from sql.utils.ding_api import DingSender
            msg_content = '''您的导出查询提取数据申请 {}：\n审核理由：{}\n实例名称：{}\n数据库：{}\n执行的sql查询：{}\n提取条数：{}\n操作时间：{}\n'''. \
                format(msg, audit_msg, ql.instance_name, ql.db_name, ql.sqllog, ql.effect_row, ql.create_time)
            DingSender().send_msg(applicant.ding_user_id, msg_content)

        qe.audit_msg = audit_msg
        qe.auditor = user
        qe.save()
    elif qe.status == 4:
        # 审核人拒绝
        msg = qe.audit_msg
    return HttpResponse(msg)


# 获取sql查询记录
@csrf_exempt
@permission_required('sql.menu_export_query', raise_exception=True)
def export_query_log(request):
    user = request.user

    limit = int(request.POST.get('limit'))
    offset = int(request.POST.get('offset'))
    limit = offset + limit

    # 获取搜索参数
    search = request.POST.get('search')
    if search is None:
        search = ''

    # 查询个人记录，超管查看所有数据
    if user.is_superuser or user.has_perm('sql.export_query_review'):
        log_count = ExportQuery.objects.filter(Q(query_log__sqllog__contains=search) |
                                                   Q(query_log__username__contains=search) |
                                                   Q(query_log__db_name__contains=search)).count()
        qe_list = ExportQuery.objects.filter(Q(query_log__sqllog__contains=search) |
                                                  Q(query_log__username__contains=search) |
                                                  Q(query_log__db_name__contains=search)).order_by('-id')[offset:limit]
    else:
        log_count = ExportQuery.objects.filter(query_log__username=user.username).filter(
                                                    Q(query_log__sqllog__contains=search) |
                                                    Q(query_log__db_name__contains=search)).count()
        qe_list = ExportQuery.objects.filter(query_log__username=user.username).filter(
                                                    Q(query_log__sqllog__contains=search) |
                                                    Q(query_log__db_name__contains=search)).order_by('-id')[offset:limit]

    sql_log_list = list()
    for qe in qe_list:
        ql = qe.query_log
        sql_log_list.append({"user_display": ql.user_display, "instance_name": ql.instance_name, "db_name": ql.db_name,
                             "create_time": ql.create_time, "sqllog": ql.sqllog, "effect_row": ql.effect_row,
                             "cost_time": ql.cost_time, "reason": qe.reason, "status": qe.status,
                             "auditor": qe.auditor.username, "id": qe.id})

    result = {"total": log_count, "rows": sql_log_list}
    # 返回查询结果
    return HttpResponse(json.dumps(result, cls=ExtendJSONEncoder, bigint_as_string=True),
                        content_type='application/json')
